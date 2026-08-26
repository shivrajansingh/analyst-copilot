"""Spreadsheets and delimited data → Markdown.

Neither format has pages, and neither should be given any. A workbook's natural
unit is the worksheet, and a delimited file's is the file itself -- those are
the locations a reader can actually navigate to. Where a sheet is too large to
embed as one unit it is cut into row blocks, and the block is labelled by the
rows it covers (`rows 502-1001`) so the citation still names something findable.

The header row is repeated at the top of every block. Without it, block three
of a revenue sheet is an anonymous grid of numbers, both to the reader and to
the embedding.
"""

from __future__ import annotations

import csv
import io
from pathlib import Path
from typing import List, Optional, Sequence

from analyst_copilot.parsing.base import DocumentParser, ParseError, ParsedSegment
from analyst_copilot.parsing.formats import DocumentFormat
from analyst_copilot.parsing.markdown import clean_cell, join_blocks, heading, render_table
from analyst_copilot.parsing.models import SegmentKind

# Rows per block. Chosen so a block lands near the embedding window rather than
# far past it: financial rows run 60-120 characters, so ~200 rows is ~15 kB of
# Markdown, which chunking downstream can still divide sensibly.
ROWS_PER_BLOCK = 200
# Below this, a sheet is one segment and gets no row-range suffix.
MIN_ROWS_TO_SPLIT = ROWS_PER_BLOCK + 50
_SNIFF_BYTES = 8192


class ExcelParser(DocumentParser):
    """One Markdown segment per worksheet, or per row block for large sheets."""

    format = DocumentFormat.XLSX
    segmentation = "worksheet"

    def segments(self, path: Path) -> List[ParsedSegment]:
        from openpyxl import load_workbook

        # `data_only` reads the cached result of a formula rather than the
        # formula text: an analyst asks what a cell says, not how it is computed.
        workbook = load_workbook(str(path), read_only=True, data_only=True)
        try:
            segments: List[ParsedSegment] = []
            for sheet in workbook.worksheets:
                rows = [
                    [clean_cell(value) for value in row]
                    for row in sheet.iter_rows(values_only=True)
                ]
                segments.extend(_sheet_segments(sheet.title, rows))
            return segments
        finally:
            workbook.close()


class CsvParser(DocumentParser):
    """A delimited file as one Markdown table, or row blocks when large."""

    format = DocumentFormat.CSV
    segmentation = "table"

    def segments(self, path: Path) -> List[ParsedSegment]:
        text = path.read_text(encoding="utf-8", errors="replace")
        if not text.strip():
            return []
        rows = [
            [clean_cell(cell) for cell in row]
            for row in csv.reader(io.StringIO(text), delimiter=_delimiter(text, path))
        ]
        if not rows:
            return []
        return _sheet_segments(path.stem, rows, noun="table", kind=SegmentKind.TABLE)


def _delimiter(text: str, path: Path) -> str:
    if path.suffix.lower() == ".tsv":
        return "\t"
    try:
        return csv.Sniffer().sniff(text[:_SNIFF_BYTES], delimiters=",;\t|").delimiter
    except csv.Error:
        # A single-column file gives the sniffer nothing to detect; comma is
        # the right reading of one column per line.
        return ","


def _sheet_segments(
    title: str,
    rows: Sequence[Sequence[str]],
    noun: str = "sheet",
    kind: SegmentKind = SegmentKind.SHEET,
) -> List[ParsedSegment]:
    populated = [list(row) for row in rows if any(cell for cell in row)]
    if not populated:
        return []

    label_stem = f"{noun} '{title}'" if title else noun

    if len(populated) < MIN_ROWS_TO_SPLIT:
        return [
            ParsedSegment(
                markdown=join_blocks([heading(label_stem, 2), render_table(populated)]),
                kind=kind,
                label=label_stem,
            )
        ]

    header, body = populated[0], populated[1:]
    segments: List[ParsedSegment] = []
    for start in range(0, len(body), ROWS_PER_BLOCK):
        block = body[start : start + ROWS_PER_BLOCK]
        # Row numbers are 1-based and count the header, matching what a
        # spreadsheet shows in its row gutter.
        first_row = start + 2
        last_row = first_row + len(block) - 1
        label = f"{label_stem} rows {first_row}-{last_row}"
        segments.append(
            ParsedSegment(
                markdown=join_blocks(
                    [heading(label, 2), render_table(block, header=header)]
                ),
                # A row block is a slice we chose, not a unit the file declares.
                kind=SegmentKind.SECTION,
                label=label,
            )
        )
    return segments


class PlainTextParser(DocumentParser):
    """Markdown and plain text, taken as-is and chunked only if oversized."""

    format = DocumentFormat.MARKDOWN
    segmentation = "whole-document"
    chars_per_section = 3500

    def __init__(self, document_format: Optional[DocumentFormat] = None) -> None:
        if document_format is not None:
            self.format = document_format

    def segments(self, path: Path) -> List[ParsedSegment]:
        try:
            text = path.read_text(encoding="utf-8", errors="replace").strip()
        except OSError as exc:
            raise ParseError(f"Could not read {path.name!r}: {exc}") from exc
        if not text:
            return []
        if len(text) <= self.chars_per_section:
            self.segmentation = "whole-document"
            return [
                ParsedSegment(markdown=text, kind=SegmentKind.SECTION, label="part 1")
            ]

        self.segmentation = "fallback-chunk"
        return [
            ParsedSegment(
                markdown=text[start : start + self.chars_per_section],
                kind=SegmentKind.SECTION,
                label=f"part {number}",
            )
            for number, start in enumerate(
                range(0, len(text), self.chars_per_section), start=1
            )
        ]
